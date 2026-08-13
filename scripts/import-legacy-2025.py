#!/usr/bin/env python3
"""Build guarded 2025 history payloads for the Engosoft dashboard.

The script is intentionally source-hash locked and does not send network
requests.  It turns the two approved workbooks into atomic Railway ingest
payloads; an operator can inspect the JSON before posting it with the ingest
secret held outside this repository.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import argparse
import hashlib
import json

from openpyxl import load_workbook


EXPECTED_ACCOUNTING_ROWS = 9_306
EXPECTED_AD_ROWS = 12_974
EXPECTED_ACCOUNTING_SHA256 = "721370182117e643bf40979cd42f7daf3ad1465421585ff35022ba11f817857f"
EXPECTED_ADS_SHA256 = "4f53515d306a064f5d37989b7e16e606ed3d0c634785372a75dbdb7885c7e389"
MAX_BYTES = 10 * 1024 * 1024


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    if isinstance(value, float):
        return format(Decimal(str(value)), "f")
    return str(value).strip()


def rows(path: Path, sheet: str):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet]
    source = worksheet.iter_rows(values_only=True)
    header = [text(value) for value in next(source)]
    try:
        for excel_row, values in enumerate(source, 2):
            yield excel_row, {
                header[index]: values[index]
                for index in range(len(header))
                if header[index]
            }
    finally:
        workbook.close()


def build_accounting(path: Path) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for excel_row, raw in rows(path, "YOY"):
        if not text(raw.get("Payment Date")).startswith("2025-"):
            continue
        movement = text(raw.get("حركة"))
        is_refund = movement.upper().startswith("RINVNT")
        row = {key: text(value) for key, value in raw.items() if key not in {"Year", "Month"}}
        # The workbook already signs $ Sales. The app signs real credit notes,
        # so store their magnitude while leaving ordinary negative discount
        # lines untouched.
        raw_usd = Decimal(text(raw.get("$ Sales")))
        source_usd = text(abs(raw_usd)) if is_refund else text(raw.get("$ Sales"))
        row.update(
            {
                "__stable_key": f"legacy-2025:yoy-row-{excel_row:05d}",
                "__legacy_source_row": str(excel_row),
                "__odoo_move_type": "out_refund" if is_refund else "out_invoice",
                "__source_usd_locked": source_usd,
                "__reporting_value_locked": "true",
            }
        )
        result.append(row)
    return result


def platform(value: str) -> str:
    normalized = value.lower()
    if "snap" in normalized:
        return "snapchat"
    if "tiktok" in normalized or "tik tok" in normalized:
        return "tiktok"
    if "facebook" in normalized or "meta" in normalized:
        return "meta"
    return ""


def build_ads(path: Path) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for excel_row, raw in rows(path, "YOY (Spend)"):
        if not text(raw.get("Date")).startswith("2025-"):
            continue
        source_platform = platform(text(raw.get("Platform")))
        row = {key: text(value) for key, value in raw.items()}
        raw_leads = text(raw.get("On-Facebook leads"))
        row.update(
            {
                "__stable_key": f"legacy-2025:yoy-spend-row-{excel_row:05d}",
                "__platform": source_platform,
                "__legacy_source_row": str(excel_row),
                "__raw_on_facebook_leads": raw_leads,
                # TikTok's historical column is not a comparable lead metric.
                # Preserve the raw cell above but fail closed for reporting.
                "Leads (Native)": "" if source_platform == "tiktok" else raw_leads,
                "__counted_platform_leads": "" if source_platform == "tiktok" else raw_leads,
            }
        )
        result.append(row)
    return result


def payload(dataset: str, data: list[dict[str, str]], source_hash: str, source: str):
    return {
        "dataset": dataset,
        "mode": "replace",
        "rows": data,
        "syncedAt": "2026-08-13T00:00:00.000Z",
        "metadata": {
            "source": source,
            "sourceSha256": source_hash,
            "period": "2025-01-01/2025-12-31",
            "rows": len(data),
            "immutableHistoricalImport": True,
        },
    }


def write_payload(path: Path, value: dict) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    assert path.stat().st_size <= MAX_BYTES, (path.name, path.stat().st_size)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--invoices", type=Path, required=True)
    parser.add_argument("--ads", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()

    assert file_hash(args.invoices) == EXPECTED_ACCOUNTING_SHA256, "Invoices workbook changed"
    assert file_hash(args.ads) == EXPECTED_ADS_SHA256, "Ads workbook changed"

    accounting = build_accounting(args.invoices)
    ads = build_ads(args.ads)
    assert len(accounting) == EXPECTED_ACCOUNTING_ROWS
    assert len(ads) == EXPECTED_AD_ROWS
    assert len({row["__stable_key"] for row in accounting}) == len(accounting)
    assert len({row["__stable_key"] for row in ads}) == len(ads)
    assert all(row["Payment Date"].startswith("2025-") for row in accounting)
    assert all(row["Date"].startswith("2025-") for row in ads)
    assert all(row["__platform"] in {"meta", "snapchat", "tiktok"} for row in ads)
    assert all(row["__counted_platform_leads"] == "" for row in ads if row["__platform"] == "tiktok")

    recognized = sum(
        -abs(Decimal(row["__source_usd_locked"]))
        if row["__odoo_move_type"] == "out_refund"
        else Decimal(row["__source_usd_locked"])
        for row in accounting
    )
    raw_revenue = sum(Decimal(row["$ Sales"]) for row in accounting)
    assert recognized == raw_revenue
    assert abs(raw_revenue - Decimal("1130492.750961334")) < Decimal("0.000001")
    assert abs(sum(Decimal(row["Cost"]) for row in ads) - Decimal("124708.76")) < Decimal("0.01")

    args.out.mkdir(parents=True, exist_ok=True)
    outputs = [
        (
            args.out / "accounting-legacy-2025.json",
            payload(
                "accounting_legacy",
                accounting,
                EXPECTED_ACCOUNTING_SHA256,
                "Invoices Analysis.xlsx!YOY",
            ),
        ),
        (
            args.out / "ads-legacy-2025.json",
            payload(
                "ads_legacy",
                ads,
                EXPECTED_ADS_SHA256,
                "Data Refrence.xlsx!YOY (Spend)",
            ),
        ),
    ]
    for output, value in outputs:
        write_payload(output, value)
        print(json.dumps({"dataset": value["dataset"], "rows": len(value["rows"]), "bytes": output.stat().st_size}))


if __name__ == "__main__":
    main()
